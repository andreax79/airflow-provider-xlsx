#!/usr/bin/env python

import os
import os.path
import datetime
import dateutil.parser
from openpyxl import load_workbook, Workbook
from airflow.exceptions import AirflowException
from airflow.models import BaseOperator
from airflow.utils.decorators import apply_defaults
from airflow_xlsx.commons import (
    clean_key,
    FileFormat,
    DEFAULT_CSV_DELIMITER,
    DEFAULT_CSV_HEADER,
    DEFAULT_FORMAT,
    HEADER_UPPER,
    HEADER_LOWER,
    XLS_EPOC,
    XLSX_EPOC,
)

__all__ = ['FromXLSXOperator']


class FromXLSXOperator(BaseOperator):
    FileFormat = FileFormat
    template_fields = ('source', 'target', 'worksheet')

    @apply_defaults
    def __init__(
        self,
        source,
        target,
        worksheet=0,
        drop_columns=None,
        add_columns=None,
        types=None,
        columns_names=None,
        file_format=DEFAULT_FORMAT,
        csv_delimiter=DEFAULT_CSV_DELIMITER,
        csv_header=DEFAULT_CSV_HEADER,
        *args,
        **kwargs
    ):
        """
        :param source: source filename (xlsx or xls)
        :param target: target filename (csv or parquet)
        :param worksheet: worksheet title or number (zero-based)
        :param drop_columns: list of columns to be dropped
        :param add_columns: columns to be added (dict or list column=value)
        :param types: force column type (dict or list column='str', 'd', 'datetime64[ns]')
        :param columns_names: force column names (list)
        :param file_format: file format ('parquet' or 'csv')
        :param csv_delimiter: CSV delimiter (defualt: ',')
        :param csv_header: CSV header case ('lower', 'upper', 'skip')
        """
        super(FromXLSXOperator, self).__init__(*args, **kwargs)
        self.source = source
        self.target = target
        try:
            self.worksheet = int(worksheet)
        except:
            self.worksheet = worksheet
        self.drop_columns = drop_columns or []
        if isinstance(add_columns, list):
            self.add_columns = dict(x.split('=') for x in add_columns)
        else:
            self.add_columns = add_columns or {}
        if isinstance(types, list):
            self.types = dict(x.split('=') for x in types)
        else:
            self.types = types or {}
        self.names = columns_names
        self.file_format = FileFormat.lookup(file_format)
        self.csv_delimiter = csv_delimiter
        self.csv_header = csv_header

    def get_type(self, name, value):
        if isinstance(value, float) or isinstance(value, int):
            return 'd'  # double
        elif isinstance(value, datetime.datetime):
            return 'datetime64[ns]'  # double
        elif isinstance(value, str):
            return 'str'
        else:
            raise Exception('unsupported data type {} {}'.format(name, type(value)))

    @classmethod
    def load_worksheet_xls(cls, filename, worksheet):
        # Load a worksheet from an XLS file
        import xlrd

        wb = xlrd.open_workbook(filename)
        if isinstance(worksheet, int):
            sheet = wb.sheets()[worksheet]
        else:  #  get by name
            t = [x for x in wb.sheet_names() if x.lower() == worksheet.lower()][0]
            if not t:
                raise Exception('Worksheet ' + worksheet + ' not found')
            sheet = t[0]
        # prepare an XLSX sheet
        xsheet = Workbook().worksheets[0]
        for row in range(0, sheet.nrows):
            for col in range(0, sheet.ncols):
                value = sheet.cell_value(row, col)
                cell_type = sheet.cell_type(row, col)
                if cell_type == xlrd.XL_CELL_DATE:
                    value = XLS_EPOC + datetime.timedelta(days=value)
                elif cell_type == xlrd.XL_CELL_NUMBER:
                    if isinstance(value, float) and value.is_integer():
                        value = int(value)
                    print(value, type(value), sheet.cell_type(row, col))
                xsheet.cell(row=row + 1, column=col + 1).value = value
        return xsheet

    @classmethod
    def load_worksheet_xlsx(cls, filename, worksheet):
        # Load a worksheet from an XLSX file
        wb = load_workbook(filename=filename, data_only=True)
        if isinstance(worksheet, int):
            sheet = wb.worksheets[worksheet]
        else:  #  get by name
            t = [x for x in wb.worksheets if x.title.lower() == worksheet.lower()][0]
            if not t:
                raise Exception('Worksheet ' + worksheet + ' not found')
            sheet = t[0]
        return sheet

    @classmethod
    def load_worksheet(cls, filename, worksheet):
        # Load a worksheet from an XLSX or XLS file
        ext = os.path.splitext(filename)[1].lower()
        if ext in ['.xls', '.xlt']:  #  Old format
            return cls.load_worksheet_xls(filename, worksheet)
        else:  #  XLS format
            return cls.load_worksheet_xlsx(filename, worksheet)

    def execute(self, context):
        try:
            sheet = self.load_worksheet(filename=self.source, worksheet=self.worksheet)
            rows = list(sheet)
            if self.names is not None:
                names = self.names
            else:
                names = [clean_key(x.value) for x in rows[0] if x.value is not None]
            datatypes = dict([(name, self.types.get(name)) for name in names])
            if '_index' in datatypes:
                datatypes['_index'] = 'd'
            columns = dict([(name, []) for name in names])
            for name, value in self.add_columns.items():
                datatypes[name] = self.get_type(name, value)
                columns[name] = []
            for _index, row in enumerate(rows[1:]):
                for i, name in enumerate(names):
                    if name == '_index':
                        columns['_index'].append(_index)
                        continue
                    cel = row[i]
                    value = cel.value
                    if isinstance(value, str):
                        value = value.strip()
                    if datatypes[name] is None and value is not None:
                        datatypes[name] = self.get_type(name, value)
                    if datatypes[name] == 'datetime64[ns]':
                        if not value:
                            value = None
                        elif isinstance(value, int) or isinstance(value, float):
                            value = XLSX_EPOC + datetime.timedelta(days=value)
                        elif not isinstance(
                            value, datetime.datetime
                        ) and not isinstance(value, datetime.date):
                            value = dateutil.parser.parse(value)
                    columns[name].append(value)
                for name, value in self.add_columns.items():
                    if name not in names:
                        columns[name].append(value)
            if self.file_format == FileFormat.csv:
                self.write_csv(columns, datatypes)
            else:
                self.write_parquet(names, columns, datatypes)
        except Exception as e:
            raise AirflowException("XLSXToParquet operator error: {0}".format(str(e)))
        return True

    def write_parquet(self, names, columns, datatypes):
        # Write the results in parquet format
        import pandas as pd
        import pyarrow.parquet

        all_names = names + [x for x in self.add_columns.keys() if x not in names]
        pd_data = {}
        for name in all_names:
            if name not in self.drop_columns:
                pd_data[name] = pd.Series(columns[name], dtype=datatypes[name])
        df = pd.DataFrame(pd_data)
        pyarrow.parquet.write_table(
            table=pyarrow.Table.from_pandas(df),
            where=self.target,
            compression='SNAPPY',
            flavor='spark',
        )

    def write_csv(self, columns, datatypes):
        # Write data to CSV file
        import csv

        data = zip(*[columns[k] for k in datatypes])
        with open(self.target, 'w') as f:
            csw_writer = csv.writer(
                f, quoting=csv.QUOTE_MINIMAL, delimiter=self.csv_delimiter
            )
            if self.csv_header == HEADER_UPPER:
                csw_writer.writerows([[x.upper() for x in datatypes.keys()]])  # header
            elif self.csv_header == HEADER_LOWER:
                csw_writer.writerows([datatypes.keys()])  # header
            csw_writer.writerows(data)


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument('filename')
    parser.add_argument('-w', '--worksheet', dest='worksheet', default=0)
    parser.add_argument('-a', '--add_col', dest='add_columns', action='append')
    parser.add_argument('-d', '--drop_col', dest='drop_columns', action='append')
    parser.add_argument('-t', '--type', dest='types', action='append')
    parser.add_argument(
        '--delimiter', dest='csv_delimiter', default=DEFAULT_CSV_DELIMITER
    )
    parser.add_argument('-o', '--output', dest='output')
    parser.add_argument(
        '--header',
        dest='csv_header',
        choices=['lower', 'upper', 'skip'],
        default=DEFAULT_CSV_HEADER,
    )
    group = parser.add_mutually_exclusive_group()
    group.add_argument('--csv', dest='file_format_csv', action='store_true')
    group.add_argument(
        '--parquet', dest='file_format_csv', action='store_false', default=False
    )
    args = parser.parse_args()
    file_format = 'csv' if args.file_format_csv else 'parquet'
    so = FromXLSXOperator(
        task_id='test',
        source=args.filename,
        target=args.output or (args.filename + '.' + file_format),
        drop_columns=args.drop_columns,
        add_columns=args.add_columns,
        types=args.types,
        file_format=file_format,
        worksheet=args.worksheet,
        csv_delimiter=args.csv_delimiter,
        csv_header=args.csv_header,
    )
    so.execute({})
