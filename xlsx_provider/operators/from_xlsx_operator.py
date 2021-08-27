#!/usr/bin/env python

import csv
import json
import os
import os.path
import datetime
import dateutil.parser
from openpyxl import load_workbook, Workbook
from airflow.exceptions import AirflowException
from airflow.models import BaseOperator
from airflow.utils.decorators import apply_defaults
from xlsx_provider.commons import (
    clean_key,
    get_type,
    col_number_to_name,
    FileFormat,
    DEFAULT_CSV_DELIMITER,
    DEFAULT_CSV_HEADER,
    DEFAULT_FORMAT,
    INDEX_COLUMN_NAME,
    HEADER_UPPER,
    HEADER_LOWER,
    XLS_EPOC,
    XLSX_EPOC,
)

__all__ = ['FromXLSXOperator']


class FromXLSXOperator(BaseOperator):
    """
    Convert an XLSX/XLS file into Parquet or CSV file

    Read an XLSX or XLS file and convert it into Parquet, CSV, JSON, JSON Lines(one line per record) file.

    :param source: source filename (XLSX or XLS, templated)
    :type source: str
    :param target: target filename (templated)
    :type target: str
    :param worksheet: worksheet title or number (zero-based, templated)
    :type worksheet: str or int
    :param drop_columns: list of columns to be dropped
    :type drop_columns: list of str
    :param add_columns: columns to be added (dict or list column=value)
    :type add_columns: list of str or dictionary of string key/value pair
    :param types: force columns types (dict or list column='str', 'd', 'datetime64[ns]')
    :type types: str or dictionary of string key/value pair
    :param columns_names: force columns names (list)
    :type columns_names: list of str
    :param limit: Row limit (default: None, templated)
    :type limit: int
    :param file_format: output file format (parquet, csv, json, jsonl)
    :type file_format: str
    :param csv_delimiter: CSV delimiter (default: ',')
    :type csv_delimiter: str
    :param csv_header: convert CSV header case ('lower', 'upper', 'skip')
    :type csv_header: str
    """

    FileFormat = FileFormat
    template_fields = ('source', 'target', 'worksheet', 'limit')
    ui_color = '#a934bd'

    @apply_defaults
    def __init__(
        self,
        source,
        target,
        worksheet=0,
        drop_columns=None,
        add_columns=None,
        types=None,
        columns_names=None,
        limit=None,
        file_format=DEFAULT_FORMAT,
        csv_delimiter=DEFAULT_CSV_DELIMITER,
        csv_header=DEFAULT_CSV_HEADER,
        *args,
        **kwargs
    ):
        super(FromXLSXOperator, self).__init__(*args, **kwargs)
        self.source = source
        self.target = target
        try:
            self.worksheet = int(worksheet)
        except:
            self.worksheet = worksheet
        self.drop_columns = drop_columns or []
        if isinstance(add_columns, list):
            self.add_columns = dict(x.split('=') for x in add_columns)
        else:
            self.add_columns = add_columns or {}
        if isinstance(types, list):
            self.types = dict(x.split('=') for x in types)
        else:
            self.types = types or {}
        self.names = columns_names
        self.file_format = FileFormat.lookup(file_format)
        self.limit = limit
        self.csv_delimiter = csv_delimiter
        self.csv_header = csv_header

    @classmethod
    def load_worksheet_xls(cls, filename, worksheet):
        # Load a worksheet from an XLS file
        import xlrd

        wb = xlrd.open_workbook(filename)
        if isinstance(worksheet, int):
            sheet = wb.sheets()[worksheet]
        else:  #  get by name
            t = [x for x in wb.sheet_names() if x.lower() == worksheet.lower()]
            if not t:
                raise KeyError('Worksheet {0} not found'.format(worksheet))
            sheet = t[0]
        # Prepare an XLSX sheet
        xsheet = Workbook().worksheets[0]
        for row in range(0, sheet.nrows):
            for col in range(0, sheet.ncols):
                value = sheet.cell_value(row, col)
                cell_type = sheet.cell_type(row, col)
                if cell_type == xlrd.XL_CELL_DATE:
                    value = XLS_EPOC + datetime.timedelta(days=value)
                elif cell_type == xlrd.XL_CELL_NUMBER:
                    if isinstance(value, float) and value.is_integer():
                        value = int(value)
                    # print(value, type(value), sheet.cell_type(row, col))
                xsheet.cell(row=row + 1, column=col + 1).value = value
        assert xsheet.max_row == sheet.nrows
        return xsheet

    @classmethod
    def load_worksheet_xlsx(cls, filename, worksheet):
        # Load a worksheet from an XLSX file
        wb = load_workbook(filename=filename, data_only=True)
        if isinstance(worksheet, int):
            sheet = wb.worksheets[worksheet]
        else:  #  get by name
            t = [x for x in wb.worksheets if x.title.lower() == worksheet.lower()]
            if not t:
                raise KeyError('Worksheet {0} not found'.format(worksheet))
            sheet = t[0]
        return sheet

    @classmethod
    def load_worksheet(cls, filename, worksheet):
        # Load a worksheet from an XLSX or XLS file
        ext = os.path.splitext(filename)[1].lower()
        if ext in ['.xls', '.xlt']:  #  Old format
            return cls.load_worksheet_xls(filename, worksheet)
        else:  #  XLS format
            return cls.load_worksheet_xlsx(filename, worksheet)

    def execute(self, context):
        try:
            sheet = self.load_worksheet(filename=self.source, worksheet=self.worksheet)
            rows = list(sheet)
            if self.names is not None:
                names = self.names
            else:
                names = [clean_key(x.value) for x in rows[0] if x.value is not None]
                # Append the column to the name if the name is not unique
                names = [
                    x
                    if (i == 0 or x not in names[: i - 1])
                    else '{}_{}'.format(x, col_number_to_name(i).lower())
                    for i, x in enumerate(names)
                ]
            # Check unique columns
            if len(set(names)) != len(names):
                duplicates = list(set([x for x in names if names.count(x) > 1]))
                raise Exception('Columns names are not unique: {0}'.format(duplicates))
            datatypes = dict([(name, self.types.get(name)) for name in names])
            if INDEX_COLUMN_NAME in datatypes:
                datatypes[INDEX_COLUMN_NAME] = 'd'
            columns = dict([(name, []) for name in names])
            for name, value in self.add_columns.items():
                datatypes[name] = get_type(name, value)
                columns[name] = []
            for _index, row in enumerate(rows[1:]):
                if self.limit is not None and _index >= self.limit:
                    break
                for i, name in enumerate(names):
                    if name == INDEX_COLUMN_NAME:
                        columns[INDEX_COLUMN_NAME].append(_index)
                        continue
                    cel = row[i]
                    value = cel.value
                    if isinstance(value, str):
                        value = value.strip()
                    if datatypes[name] is None and value is not None:
                        datatypes[name] = get_type(name, value)
                    if datatypes[name] == 'datetime64[ns]':
                        if not value:
                            value = None
                        elif isinstance(value, int) or isinstance(value, float):
                            value = XLSX_EPOC + datetime.timedelta(days=value)
                        elif not isinstance(
                            value, datetime.datetime
                        ) and not isinstance(value, datetime.date):
                            value = dateutil.parser.parse(value)
                    columns[name].append(value)
                for name, value in self.add_columns.items():
                    if name not in names:
                        columns[name].append(value)
            row_num = sheet.max_row - 1
            if self.limit is not None:
                row_num = min(row_num, self.limit)
            for i, name in enumerate(names):
                assert len(columns[name]) == row_num  # check rows number (skip header)
            self.write(names, columns, datatypes)
        except Exception as e:
            raise AirflowException("XLSXToParquet operator error: {0}".format(str(e)))
        return True

    def write_parquet(self, names, columns, datatypes):
        "Write the results in parquet format"
        import pandas as pd
        import pyarrow.parquet

        all_names = names + [x for x in self.add_columns.keys() if x not in names]
        pd_data = {}
        for name in all_names:
            if name not in self.drop_columns:
                pd_data[name] = pd.Series(columns[name], dtype=datatypes[name])
        df = pd.DataFrame(pd_data)
        pyarrow.parquet.write_table(
            table=pyarrow.Table.from_pandas(df),
            where=self.target,
            compression='SNAPPY',
            flavor='spark',
        )

    def write_csv(self, names, columns, datatypes):
        "Write data to CSV file"
        data = zip(*[columns[k] for k in datatypes])
        with open(self.target, 'w') as f:
            csw_writer = csv.writer(
                f, quoting=csv.QUOTE_MINIMAL, delimiter=self.csv_delimiter
            )
            if self.csv_header == HEADER_UPPER:
                csw_writer.writerows([[x.upper() for x in datatypes.keys()]])
            elif self.csv_header == HEADER_LOWER:
                csw_writer.writerows([datatypes.keys()])
            csw_writer.writerows(data)

    def write_json(self, names, columns, datatypes):
        "Write data to JSON file"
        data = list(
            dict(q) for q in zip(*(list((k, x) for x in v) for k, v in columns.items()))
        )
        with open(self.target, 'w') as f:
            f.write(json.dumps(data, indent=2, default=str))

    def write_jsonl(self, names, columns, datatypes):
        "Write data to JSON Lines file"
        data = list(
            dict(q) for q in zip(*(list((k, x) for x in v) for k, v in columns.items()))
        )
        with open(self.target, 'w') as f:
            f.write('\n'.join(json.dumps(x, default=str) for x in data))

    def write(self, names, columns, datatypes):
        "Write data to file"
        if self.file_format == FileFormat.csv:
            self.write_csv(names, columns, datatypes)
        elif self.file_format == FileFormat.json:
            self.write_json(names, columns, datatypes)
        elif self.file_format == FileFormat.jsonl:
            self.write_jsonl(names, columns, datatypes)
        else:
            self.write_parquet(names, columns, datatypes)


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument('filename')
    parser.add_argument('-w', '--worksheet', dest='worksheet', default=0)
    parser.add_argument('-a', '--add_col', dest='add_columns', action='append')
    parser.add_argument('-d', '--drop_col', dest='drop_columns', action='append')
    parser.add_argument('-t', '--type', dest='types', action='append')
    parser.add_argument(
        '--delimiter', dest='csv_delimiter', default=DEFAULT_CSV_DELIMITER
    )
    parser.add_argument('-o', '--output', dest='output')
    parser.add_argument(
        '--header',
        dest='csv_header',
        choices=['lower', 'upper', 'skip'],
        default=DEFAULT_CSV_HEADER,
    )
    group = parser.add_mutually_exclusive_group()
    group.add_argument('--csv', dest='file_format_csv', action='store_true')
    group.add_argument(
        '--parquet', dest='file_format_csv', action='store_false', default=False
    )
    args = parser.parse_args()
    file_format = 'csv' if args.file_format_csv else 'parquet'
    so = FromXLSXOperator(
        task_id='test',
        source=args.filename,
        target=args.output or (args.filename + '.' + file_format),
        drop_columns=args.drop_columns,
        add_columns=args.add_columns,
        types=args.types,
        file_format=file_format,
        worksheet=args.worksheet,
        csv_delimiter=args.csv_delimiter,
        csv_header=args.csv_header,
    )
    so.execute({})
