#!/usr/bin/env python

import os
import os.path
import csv
import json
import datetime
from openpyxl import load_workbook, Workbook
from xlsx_provider.commons import copy_cells, DEFAULT_CSV_DELIMITER, XLS_EPOC

__all__ = ['load_worksheet']

loaders = {}


def extension(*exts):
    "File extensions decorator"

    def wrap(f):
        for ext in exts:
            loaders[ext] = f
        return f

    return wrap


@extension('csv')
def load_worksheet_csv(filename, sheet=None, skip_rows=0, csv_delimiter=None, **kwargs):
    "Load a worksheet from a CSV file"
    if sheet is None:
        sheet = Workbook().active  # Create a new Workbook and get the active sheet
    with open(filename, 'r', encoding='utf8') as f:
        reader = csv.reader(f, delimiter=csv_delimiter)
        for i, row in enumerate(reader):
            if i >= skip_rows:
                sheet.append(row)
    return sheet


@extension('xls', 'xlt')
def load_worksheet_xls(filename, sheet=None, skip_rows=0, worksheet=0, **kwargs):
    "Load a worksheet from an XLS file"
    import xlrd

    if sheet is None:
        sheet = Workbook().active  # Create a new Workbook and get the active sheet
    wb = xlrd.open_workbook(filename)
    if isinstance(worksheet, int):
        xsheet = wb.sheets()[worksheet]
    else:  #  get by name
        t = [x for x in wb.sheet_names() if x.lower() == worksheet.lower()]
        if not t:
            raise KeyError('Worksheet {0} not found'.format(worksheet))
        xsheet = t[0]
    # Prepare an XLSX sheet
    for row in range(skip_rows, xsheet.nrows):
        for col in range(0, xsheet.ncols):
            value = xsheet.cell_value(row, col)
            cell_type = xsheet.cell_type(row, col)
            if cell_type == xlrd.XL_CELL_DATE:
                value = XLS_EPOC + datetime.timedelta(days=value)
            elif cell_type == xlrd.XL_CELL_NUMBER:
                if isinstance(value, float) and value.is_integer():
                    value = int(value)
                # print(value, type(value), xsheet.cell_type(row, col))
            sheet.cell(row=row + 1, column=col + 1).value = value
    assert sheet.max_row == xsheet.nrows
    return sheet


@extension('xlsx', 'xlsm', 'xlsb')
def load_worksheet_xlsx(filename, sheet=None, skip_rows=0, worksheet=0, **kwargs):
    "Load a worksheet from an XLSX file"
    wb = load_workbook(filename=filename, data_only=True)
    if isinstance(worksheet, int):
        xsheet = wb.worksheets[worksheet]
    else:  #  get by name
        t = [x for x in wb.worksheets if x.title.lower() == worksheet.lower()]
        if not t:
            raise KeyError('Worksheet {0} not found'.format(worksheet))
        xsheet = t[0]
    if skip_rows:
        xsheet.delete_rows(idx=1, amount=skip_rows)
    if sheet is None:
        return xsheet
    else:
        copy_cells(xsheet, sheet)
        return sheet


@extension('parquet')
def read_parquet(filename, sheet=None, skip_rows=0, **kwargs):
    "Load a worksheet from a Parquet file"
    import pandas as pd

    if sheet is None:
        sheet = Workbook().active  # Create a new Workbook and get the active sheet
    f = pd.read_parquet(filename)
    # header
    sheet.append(list(f.columns))
    # rows
    for i, row in enumerate(f.values):
        if i >= skip_rows:
            sheet.append(list(row))
    return sheet


@extension('json')
def read_json(filename, sheet=None, skip_rows=0, **kwargs):
    "Load a worksheet from a JSON file"
    if sheet is None:
        wb = Workbook()
        sheet = wb.active
    wb = Workbook()
    with open(filename, 'r', encoding='utf8') as f:
        try:
            data = json.load(f)
        except json.decoder.JSONDecodeError as ex:
            # Try load as JSON Lines
            try:
                f.seek(0)  # rewind
                data = json.loads('[' + f.read().replace('\n', ',') + ']')
            except json.decoder.JSONDecodeError:
                raise ex
        keys = list(data[0].keys())
        # header
        sheet.append(keys)
        # rows
        for i, row in enumerate(data):
            if i >= skip_rows:
                sheet.append([row.get(key) for key in keys])
    return sheet


@extension('jsonl')
def read_jsonl(filename, sheet=None, skip_rows=0, **kwargs):
    "Load a worksheet from a JSON Lines file"
    if sheet is None:
        sheet = Workbook().active  # Create a new Workbook and get the active sheet
    with open(filename, 'r', encoding='utf8') as f:
        data = json.loads('[' + f.read().replace('\n', ',') + ']')
        keys = list(data[0].keys())
        # header
        sheet.append(keys)
        # rows
        for i, row in enumerate(data):
            if i >= skip_rows:
                sheet.append([row.get(key) for key in keys])
    return sheet


def load_worksheet(
    filename,
    sheet=None,
    worksheet=0,
    skip_rows=0,
    csv_delimiter=DEFAULT_CSV_DELIMITER,
    ext=None,
):
    """
    Load a worksheet from a supported file format

    :param filename: File to be loaded
    :type filename: str
    :param sheet: If not None, load the data into the sheet
    :type sheet: Worksheet
    :param worksheet: Worksheet title or number (zero-based)
    :type worksheet: str or int
    :param skip_rows: Number of input lines to skip
    :type skip_rows: int
    :param csv_delimiter: CSV delimiter
    :type csv_delimiter: str
    :param ext: Force file format (autodetect by default)
    :type ext: str
    """
    if ext is None:
        ext = os.path.splitext(filename)[1]
    ext = ext.lower().lstrip('.')
    loader = loaders.get(ext)
    if loader is None:
        raise KeyError('unsupported file format {}'.format(ext))
    return loader(
        filename=filename,
        sheet=sheet,
        skip_rows=skip_rows,
        worksheet=worksheet,
        csv_delimiter=csv_delimiter,
    )
